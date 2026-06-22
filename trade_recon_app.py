"""
================================================================================
 TRADE RECONCILIATION UTILITY v2  -  ID-first matching engine
================================================================================
Run with:
    streamlit run trade_recon_app.py

WHY V2 EXISTS
-------------
The first version of this tool grouped trades by business attributes
(instrument + direction + date + price) and only used Trade ID as an
afterthought. Real-world usage showed this was the wrong model: the broker
file and the Murex extract are LINKED BY A SHARED UNIQUE ID on every trade
(broker's "Unique Trade ID" / "Orig ID" == Murex's "G.ID", once a system
prefix like "BFU:" or "ICE:" is stripped). The manual process described by
the business is, in order:

  1. Trade-count check        : row counts in each file (informational).
  2. ID-existence check       : every broker ID must appear in Murex (and
                                 vice versa). Missing IDs = clear exceptions.
  3. Field-level recon per ID : for each ID present in both files, compare
                                 direction, price, and lots/nominal.
       - If broker has MULTIPLE rows for the same ID (multi-leg trade),
         their lots are SUMMED first (price is expected to be identical
         across legs of the same ID) before comparing to Murex's single
         aggregated row for that ID.
       - PASS only if direction matches AND price matches AND
         sum(broker lots for ID) == Murex nominal for ID.
       - Otherwise: FAIL, with the exact field(s), broker value(s), and
         Murex value(s) called out.

This file rebuilds the engine around that exact logic. Attribute-based
grouping/fuzzy-matching is kept ONLY as a fallback for the (rare) case
where a trade genuinely has no usable ID on one side - it is no longer the
primary mechanism.

ALL THE "POINTERS" (THINGS YOU MAY NEED TO CHANGE) ARE IN SECTION 1
---------------------------------------------------------------------
1. CANONICAL_FIELDS               -> the common field set both files map onto
2. DEFAULT_BROKER_COLUMN_MAP        -> Broker file column -> canonical field
3. DEFAULT_MUREX_COLUMN_MAP         -> Murex file column -> canonical field
4. ID_STRIP_PREFIXES                -> system prefixes to strip from IDs
                                        (e.g. "BFU:", "ICE:", "GID:")
5. DIRECTION_BUY_VALUES / SELL_VALUES
6. Tolerances (price / lots / date) -> editable in the UI too
================================================================================
"""

import re
from dataclasses import dataclass, field
from typing import Dict, List, Optional, Tuple
import tempfile

import pandas as pd
import streamlit as st
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ==============================================================================
# 1. CONFIGURATION  -  edit this section to point the tool at your files
# ==============================================================================

# --- 1a. Canonical schema ----------------------------------------------------
CANONICAL_FIELDS = [
    "trade_id",       # the system's own row/leg identifier (Deal ID, TRN.NB...)
    "link_id",        # THE shared reconciliation key (Broker "Unique Trade ID" /
                       # "Orig ID"  <->  Murex "G.ID"). This is the PRIMARY
                       # matching field in this design.
    "instrument",      # product / instrument code
    "direction",        # normalised to "BUY" / "SELL"
    "trade_date",        # execution / transaction date
    "prompt_date",        # delivery / maturity / prompt month
    "price",               # traded price / rate
    "quantity",            # lots / nominal
    "currency",
    "broker",
    "counterparty",
    "trader",         # trader name. Optional - only used by the cross-ID
                       # aggregation enhancement layer (Section 7). Leaving
                       # this unmapped fully preserves all existing behaviour.
]

# --- 1b. Default column mapping: Broker (SFTP) file --------------------------
# Matches an ICE/Trafigura-style SFTP export, and also the
# "Unique Trade ID / Lots / Price" style file shown in the manual process doc.
DEFAULT_BROKER_COLUMN_MAP: Dict[str, Optional[str]] = {
    "trade_id":     "Deal ID",
    "link_id":      "Orig ID",          # falls back to "Unique Trade ID" if present instead
    "instrument":   "Product",
    "direction":    "B/S",
    "trade_date":   "Trade Date",
    "prompt_date":  "Contract",
    "price":        "Price",
    "quantity":     "Lots",
    "currency":     None,
    "broker":       "Clearing Firm",
    "counterparty": "Trading Company",
    "trader":       "Trader",
}

# --- 1c. Default column mapping: Murex extract --------------------------------
# Matches a Murex Trade Query export (TRN.NB / G.ID / B/S / RATE / NOMINAL...).
DEFAULT_MUREX_COLUMN_MAP: Dict[str, Optional[str]] = {
    "trade_id":     "TRN.NB",
    "link_id":      "G.ID",
    "instrument":   "PL INSTRUMENT",
    "direction":    "B/S",
    "trade_date":   "TRN.DATE",
    "prompt_date":  "START",
    "price":        "RATE",
    "quantity":     "NOMINAL",
    "currency":     "PL CUR",
    "broker":       "BRK_CTP",
    "counterparty": "CTP/PFL 2",
    "trader":       "TRADER",
}

# --- 1d. ID prefixes to strip when normalising link_id -----------------------
# Murex G.ID values look like "BFU:75287073" or "ICE:922776111" - the part
# before ":" is a SOURCE-SYSTEM TAG, not part of the trade identifier, and
# must be stripped before comparing to the broker's plain numeric ID.
ID_STRIP_PREFIXES = ["BFU", "ICE", "GID", "BBG", "TPI"]

# --- 1e. Direction normalisation ----------------------------------------------
DIRECTION_BUY_VALUES = {"B", "BUY", "BOT", "BOUGHT", "BUYER"}
DIRECTION_SELL_VALUES = {"S", "SELL", "SOLD", "SELLER"}

# --- 1f. Tolerances & fallback matching ---------------------------------------
DEFAULT_PRICE_TOLERANCE = 0.0001      # absolute price/rate difference allowed
DEFAULT_QTY_TOLERANCE = 0.0           # absolute lots/nominal difference allowed
DEFAULT_GROUP_KEYS = ["instrument", "direction", "prompt_date", "trade_date", "price"]
DEFAULT_ENABLE_FALLBACK_MATCHING = True   # attribute-based matching for no-ID trades
SUBSET_SUM_ITEM_LIMIT = 60

# --- 1g. Cross-ID aggregation (enhancement layer, Section 7) ------------------
# This is an OPTIONAL, ADDITIVE post-processing step that runs only after the
# normal ID-based reconciliation above has already produced its result. It
# never changes how PASS/FAIL/Missing are originally decided - it only ever
# looks at ID-matched pairs that FAILED specifically on quantity, and asks:
# "do other orphaned broker rows (different IDs entirely), filtered by
# trader + price (+ commodity if still ambiguous), sum exactly to the Murex
# quantity?" If yes - and ONLY if there is exactly one such combination, with
# no ambiguity - the case is moved out of FAIL/Missing into a separate
# "Resolved by Cross-ID Aggregation" category. Ties (more than one valid
# combination) are deliberately left as FAIL for manual review - this layer
# never guesses. Turning DEFAULT_ENABLE_CROSS_ID_AGGREGATION off (or
# unchecking it in the UI) fully restores the utility's original behaviour.
DEFAULT_ENABLE_CROSS_ID_AGGREGATION = True
CROSS_ID_PRICE_TOLERANCE = 0.0          # exact match only, per business sign-off
CROSS_ID_SUBSET_SUM_ITEM_LIMIT = 200    # no practical cap requested; kept as a
                                         # safety ceiling against pathological input

# --- 1h. Missing-set attribute aggregation (new enhancement layer, Section 8) ------
# Runs AFTER the main reconcile() and AFTER the cross-ID aggregation in 3B.
# For Murex IDs that land in Missing-in-Broker (no direct broker ID match at all),
# searches among broker rows sitting in Missing-in-Murex for a unique combination
# whose lots sum equals the Murex quantity, filtered by price (exact), then trader,
# then commodity if still ambiguous.  Same "no guessing" policy as cross-ID.
DEFAULT_ENABLE_MISSING_ATTR_RESOLUTION = True


@dataclass
class ReconConfig:
    price_tolerance: float = DEFAULT_PRICE_TOLERANCE
    quantity_tolerance: float = DEFAULT_QTY_TOLERANCE
    group_keys: List[str] = field(default_factory=lambda: list(DEFAULT_GROUP_KEYS))
    enable_fallback_matching: bool = DEFAULT_ENABLE_FALLBACK_MATCHING
    subset_sum_item_limit: int = SUBSET_SUM_ITEM_LIMIT
    enable_cross_id_aggregation: bool = DEFAULT_ENABLE_CROSS_ID_AGGREGATION
    cross_id_subset_sum_item_limit: int = CROSS_ID_SUBSET_SUM_ITEM_LIMIT
    enable_missing_attr_resolution: bool = DEFAULT_ENABLE_MISSING_ATTR_RESOLUTION


# ==============================================================================
# 2. DATA LOADING & NORMALISATION
# ==============================================================================

_MONTH_MAP = {
    "JAN": 1, "FEB": 2, "MAR": 3, "APR": 4, "MAY": 5, "JUN": 6,
    "JUL": 7, "AUG": 8, "SEP": 9, "OCT": 10, "NOV": 11, "DEC": 12,
}


def read_any(file_obj) -> pd.DataFrame:
    name = getattr(file_obj, "name", "")
    if str(name).lower().endswith(".csv"):
        return pd.read_csv(file_obj, dtype=str, keep_default_na=False)
    return pd.read_excel(file_obj, dtype=str)


def _normalise_link_id(value, strip_prefixes: List[str]) -> Optional[str]:
    """Strip known system prefixes (e.g. 'BFU:', 'ICE:') and leading zeros so
    a Murex G.ID lines up exactly with the broker's plain trade ID."""
    if value is None or str(value).strip() == "":
        return None
    v = str(value).strip()
    for sep in (":", "_", "-"):
        if sep in v:
            head, _, tail = v.partition(sep)
            if head.upper() in [p.upper() for p in strip_prefixes]:
                v = tail.strip()
                break
    if v.isdigit():
        return str(int(v))  # drop leading zeros for safe comparison
    return v.upper() or None


def _normalise_direction(value):
    if value is None:
        return None
    v = str(value).strip().upper()
    if v in DIRECTION_BUY_VALUES:
        return "BUY"
    if v in DIRECTION_SELL_VALUES:
        return "SELL"
    return v or None


def _normalise_price(value):
    if value is None or str(value).strip() == "":
        return None
    try:
        return float(str(value).replace(",", "").strip())
    except ValueError:
        return None


def _normalise_quantity(value):
    if value is None or str(value).strip() == "":
        return 0.0
    try:
        return float(str(value).replace(",", "").strip())
    except ValueError:
        return 0.0


def _normalise_trade_date(value):
    if value is None or str(value).strip() == "":
        return None
    v = str(value).strip()
    if re.fullmatch(r"\d{8}", v):
        return pd.to_datetime(v, format="%Y%m%d", errors="coerce")
    if re.fullmatch(r"\d{4}-\d{2}-\d{2}", v):
        return pd.to_datetime(v, format="%Y-%m-%d", errors="coerce")
    if re.fullmatch(r"\d{4}-\d{2}-\d{2}[ T]\d{2}:\d{2}:\d{2}", v):
        return pd.to_datetime(v, format="%Y-%m-%d %H:%M:%S", errors="coerce")
    return pd.to_datetime(v, errors="coerce", dayfirst=True)


def _normalise_prompt_date(value):
    if value is None or str(value).strip() == "":
        return None, None
    v = str(value).strip()
    m = re.fullmatch(r"([A-Za-z]{3})-?(\d{2,4})", v)
    if m:
        mon = _MONTH_MAP.get(m.group(1).upper())
        yr = int(m.group(2))
        yr = yr + 2000 if yr < 100 else yr
        if mon:
            return (yr, mon), f"{m.group(1).title()}-{yr % 100:02d}"
    if re.fullmatch(r"\d{8}", v):
        ts = pd.to_datetime(v, format="%Y%m%d", errors="coerce")
        if pd.notna(ts):
            return (ts.year, ts.month), ts.strftime("%b-%y")
    ts = pd.to_datetime(v, errors="coerce", dayfirst=True)
    if pd.notna(ts):
        return (ts.year, ts.month), ts.strftime("%b-%y")
    return (None, None), v


def normalise(df: pd.DataFrame, column_map: dict, source_label: str,
              id_strip_prefixes: List[str]) -> pd.DataFrame:
    out = pd.DataFrame(index=df.index)
    for fname in CANONICAL_FIELDS:
        src_col = column_map.get(fname)
        out[fname] = df[src_col] if (src_col and src_col in df.columns) else None

    out["direction"] = out["direction"].map(_normalise_direction)
    out["price"] = out["price"].map(_normalise_price)
    out["quantity"] = out["quantity"].map(_normalise_quantity)
    out["trade_date"] = out["trade_date"].map(_normalise_trade_date)
    out["link_id"] = out["link_id"].map(lambda v: _normalise_link_id(v, id_strip_prefixes))

    prompt_norm = out["prompt_date"].map(_normalise_prompt_date)
    out["prompt_date_key"] = prompt_norm.map(lambda x: x[0])
    out["prompt_date"] = prompt_norm.map(lambda x: x[1])

    out["instrument"] = out["instrument"].astype(str).str.strip().str.upper()
    out["instrument"] = out["instrument"].replace({"NONE": None, "NAN": None})

    for col in ("trade_id", "currency", "broker", "counterparty"):
        out[col] = out[col].astype(str).str.strip()
        out[col] = out[col].replace({"None": None, "nan": None, "": None})

    # Normalise trader to uppercase so matching is case-insensitive across files.
    out["trader"] = out["trader"].astype(str).str.strip().str.upper()
    out["trader"] = out["trader"].replace({"NONE": None, "NAN": None, "": None})

    out["_source"] = source_label
    out["_row_ref"] = df.index + 2
    return out


# ==============================================================================
# 3. RECONCILIATION ENGINE  -  ID-first design
# ==============================================================================

def _fmt(row: pd.Series) -> dict:
    out = {}
    for f in ("trade_id", "link_id", "instrument", "direction", "trade_date",
              "prompt_date", "price", "quantity", "currency", "broker",
              "counterparty", "trader", "_row_ref"):
        v = row.get(f)
        if isinstance(v, pd.Timestamp):
            v = v.strftime("%Y-%m-%d") if pd.notna(v) else None
        out[f] = v
    return out


def _fmt_many(df: pd.DataFrame, idxs) -> List[dict]:
    return [_fmt(df.loc[i]) for i in idxs]


@dataclass
class ReconResult:
    id_passed: List[dict] = field(default_factory=list)        # PASS, single leg
    id_passed_aggregated: List[dict] = field(default_factory=list)  # PASS, multi-leg rollup
    id_failed: List[dict] = field(default_factory=list)         # ID present both sides, field mismatch
    missing_in_murex: List[dict] = field(default_factory=list)  # ID only in broker
    missing_in_broker: List[dict] = field(default_factory=list) # ID only in murex
    fallback_matched: List[dict] = field(default_factory=list)  # no-ID fallback matches
    fallback_aggregated: List[dict] = field(default_factory=list)
    no_id_unmatched_broker: List[dict] = field(default_factory=list)
    no_id_unmatched_murex: List[dict] = field(default_factory=list)
    cross_id_resolved: List[dict] = field(default_factory=list)  # Enhancement layer
                                                                    # (Section 7): cases
                                                                    # resolved by combining
                                                                    # orphaned broker rows
                                                                    # across DIFFERENT IDs.
    missing_attr_resolved: List[dict] = field(default_factory=list) # Enhancement layer
                                                                     # (Section 8): Murex
                                                                     # orphans resolved from
                                                                     # broker orphan pool by
                                                                     # trader+price+commodity
                                                                     # lots-sum match.
    broker_count: int = 0
    murex_count: int = 0
    broker_ids_count: int = 0
    murex_ids_count: int = 0


def _compare_id_group(link_id: str, b_rows: pd.DataFrame, m_rows: pd.DataFrame,
                       config: ReconConfig) -> dict:
    """Compare ALL broker rows and ALL murex rows sharing one link_id.
    Per the business process: broker legs for the same ID are summed (lots)
    before comparison; price/direction are expected to be identical across
    legs and are checked for consistency too."""

    b_qty_total = b_rows["quantity"].sum()
    m_qty_total = m_rows["quantity"].sum()

    # Within-broker-side consistency check (legs of the same ID should agree
    # on price & direction - if they don't, that's its own exception)
    b_prices = b_rows["price"].dropna().unique()
    b_directions = b_rows["direction"].dropna().unique()
    m_prices = m_rows["price"].dropna().unique()
    m_directions = m_rows["direction"].dropna().unique()

    issues = []

    if len(b_directions) > 1:
        issues.append({"field": "direction", "reason": "Broker legs disagree on direction for this ID",
                        "broker_value": ", ".join(b_directions), "murex_value": ""})
    if len(m_directions) > 1:
        issues.append({"field": "direction", "reason": "Murex legs disagree on direction for this ID",
                        "broker_value": "", "murex_value": ", ".join(m_directions)})
    if len(b_directions) == 1 and len(m_directions) == 1 and b_directions[0] != m_directions[0]:
        issues.append({"field": "direction", "reason": "Direction mismatch",
                        "broker_value": b_directions[0], "murex_value": m_directions[0]})

    if len(b_prices) > 1:
        issues.append({"field": "price", "reason": "Broker legs have inconsistent price for this ID",
                        "broker_value": ", ".join(str(p) for p in b_prices), "murex_value": ""})
    if len(m_prices) > 1:
        issues.append({"field": "price", "reason": "Murex legs have inconsistent price for this ID",
                        "broker_value": "", "murex_value": ", ".join(str(p) for p in m_prices)})
    if len(b_prices) == 1 and len(m_prices) == 1:
        if abs(b_prices[0] - m_prices[0]) > config.price_tolerance:
            issues.append({"field": "price", "reason": "Price mismatch",
                            "broker_value": b_prices[0], "murex_value": m_prices[0]})

    if abs(b_qty_total - m_qty_total) > config.quantity_tolerance:
        issues.append({"field": "quantity", "reason": "Lots/Quantity mismatch (broker total vs Murex total)",
                        "broker_value": b_qty_total, "murex_value": m_qty_total})

    # trade_date sanity check (informational - included in failure detail only
    # if everything else matches but the date is off, since date is part of
    # the search criteria used to pull the Murex extract in the first place)
    b_dates = b_rows["trade_date"].dropna().unique()
    m_dates = m_rows["trade_date"].dropna().unique()
    if len(b_dates) == 1 and len(m_dates) == 1 and pd.Timestamp(b_dates[0]).date() != pd.Timestamp(m_dates[0]).date():
        issues.append({"field": "trade_date", "reason": "Trade date mismatch",
                        "broker_value": pd.Timestamp(b_dates[0]).strftime("%Y-%m-%d"),
                        "murex_value": pd.Timestamp(m_dates[0]).strftime("%Y-%m-%d")})

    is_aggregated = len(b_rows) > 1 or len(m_rows) > 1
    record = {
        "link_id": link_id,
        "is_aggregated": is_aggregated,
        "broker_legs": _fmt_many(b_rows, b_rows.index),
        "murex_legs": _fmt_many(m_rows, m_rows.index),
        "broker_qty_total": b_qty_total,
        "murex_qty_total": m_qty_total,
        "broker_price": b_prices[0] if len(b_prices) == 1 else None,
        "murex_price": m_prices[0] if len(m_prices) == 1 else None,
        "broker_direction": b_directions[0] if len(b_directions) == 1 else None,
        "murex_direction": m_directions[0] if len(m_directions) == 1 else None,
    }

    if issues:
        record["issues"] = issues
        return {"status": "FAIL", "record": record}
    return {"status": "PASS", "record": record}


# ---------------------------------------------------------------------------
# Fallback attribute-based matching (only for rows with no usable link_id)
# ---------------------------------------------------------------------------

def _group_key(row, group_keys) -> tuple:
    key = []
    for k in group_keys:
        if k == "prompt_date":
            key.append(row.get("prompt_date_key"))
        elif k == "price":
            v = row.get("price")
            key.append(round(v, 4) if v is not None and pd.notna(v) else None)
        elif k == "trade_date":
            v = row.get("trade_date")
            key.append(v.date() if pd.notna(v) else None)
        else:
            key.append(row.get(k))
    return tuple(key)


def _scale_int(value: float, scale: int = 100) -> int:
    return int(round(float(value) * scale))


def _find_subset_sum(items, target: float, tol: float, scale: int = 100):
    target_i = _scale_int(target, scale)
    tol_i = max(0, _scale_int(tol, scale))
    dp = {0: []}
    cap = target_i + tol_i
    for item_id, qty in items:
        q = _scale_int(qty, scale)
        if q <= 0:
            continue
        for s in list(dp.keys()):
            ns = s + q
            if ns <= cap and ns not in dp:
                dp[ns] = dp[s] + [item_id]
    best = None
    for s, ids in dp.items():
        if not ids:
            continue
        if abs(s - target_i) <= tol_i:
            if best is None or abs(s - target_i) < abs(best[0] - target_i):
                best = (s, ids)
    return best[1] if best else None


def _partition_many_to_many(b_items, m_items, config: ReconConfig):
    tol = max(config.quantity_tolerance, 0.001)
    if len(b_items) <= config.subset_sum_item_limit:
        remaining = list(b_items)
        assignment = {}
        for m_id, m_qty in sorted(m_items, key=lambda x: -x[1]):
            subset_ids = _find_subset_sum(remaining, m_qty, tol)
            if subset_ids is None:
                return None
            assignment[m_id] = subset_ids
            id_set = set(subset_ids)
            remaining = [it for it in remaining if it[0] not in id_set]
        if remaining:
            return None
        return assignment
    remaining = sorted(b_items, key=lambda x: -x[1])
    assignment = {}
    for m_id, m_qty in sorted(m_items, key=lambda x: -x[1]):
        bucket, total, idx = [], 0.0, 0
        while idx < len(remaining) and total < m_qty - tol:
            bid, bqty = remaining[idx]
            if total + bqty <= m_qty + tol:
                bucket.append(bid); total += bqty; remaining.pop(idx)
            else:
                idx += 1
        if abs(total - m_qty) > tol:
            return None
        assignment[m_id] = bucket
    if remaining:
        return None
    return assignment


def _fallback_attribute_match(b_df: pd.DataFrame, m_df: pd.DataFrame, config: ReconConfig,
                               result: ReconResult):
    """Attribute-based matching for the subset of rows that had NO usable
    link_id on one or both sides. Mirrors the v1 grouping logic, kept only
    as a safety net - in normal operation this should rarely be needed
    since the business process relies on the shared ID."""
    if not config.enable_fallback_matching or (b_df.empty and m_df.empty):
        result.no_id_unmatched_broker += _fmt_many(b_df, b_df.index)
        result.no_id_unmatched_murex += _fmt_many(m_df, m_df.index)
        return

    b_df = b_df.copy()
    m_df = m_df.copy()
    b_df["_gk"] = b_df.apply(lambda r: _group_key(r, config.group_keys), axis=1)
    m_df["_gk"] = m_df.apply(lambda r: _group_key(r, config.group_keys), axis=1)

    all_keys = set(b_df["_gk"]) | set(m_df["_gk"])
    for key in all_keys:
        b_sub = b_df[b_df["_gk"] == key]
        m_sub = m_df[m_df["_gk"] == key]
        b_rem, m_rem = list(b_sub.index), list(m_sub.index)

        # 1:1 exact qty match
        for b_idx in list(b_rem):
            b_qty = b_sub.at[b_idx, "quantity"]
            best_m, best_diff = None, None
            for m_idx in m_rem:
                diff = abs(b_qty - m_sub.at[m_idx, "quantity"])
                if diff <= config.quantity_tolerance and (best_diff is None or diff < best_diff):
                    best_m, best_diff = m_idx, diff
            if best_m is not None:
                result.fallback_matched.append({
                    "broker": _fmt(b_sub.loc[b_idx]), "murex": _fmt(m_sub.loc[best_m]),
                })
                b_rem.remove(b_idx); m_rem.remove(best_m)

        # aggregation match
        if b_rem and m_rem:
            sum_b = sum(b_sub.at[i, "quantity"] for i in b_rem)
            sum_m = sum(m_sub.at[i, "quantity"] for i in m_rem)
            tol = max(config.quantity_tolerance, 0.001)
            if abs(sum_b - sum_m) <= tol and (len(b_rem) > 1 or len(m_rem) > 1):
                b_items = [(i, b_sub.at[i, "quantity"]) for i in b_rem]
                m_items = [(i, m_sub.at[i, "quantity"]) for i in m_rem]
                assignment = _partition_many_to_many(b_items, m_items, config)
                if assignment:
                    for m_id, b_ids in assignment.items():
                        result.fallback_aggregated.append({
                            "broker_trades": _fmt_many(b_sub, b_ids),
                            "murex_trades": _fmt_many(m_sub, [m_id]),
                            "broker_qty_total": sum(b_sub.at[i, "quantity"] for i in b_ids),
                            "murex_qty_total": m_sub.at[m_id, "quantity"],
                        })
                    b_rem, m_rem = [], []

        result.no_id_unmatched_broker += _fmt_many(b_sub, b_rem)
        result.no_id_unmatched_murex += _fmt_many(m_sub, m_rem)


# ---------------------------------------------------------------------------
# Top-level entry point
# ---------------------------------------------------------------------------

def reconcile(broker_df: pd.DataFrame, murex_df: pd.DataFrame, config: ReconConfig) -> ReconResult:
    result = ReconResult(broker_count=len(broker_df), murex_count=len(murex_df))

    has_id_b = broker_df["link_id"].notna()
    has_id_m = murex_df["link_id"].notna()

    b_with_id = broker_df[has_id_b]
    m_with_id = murex_df[has_id_m]
    b_without_id = broker_df[~has_id_b]
    m_without_id = murex_df[~has_id_m]

    b_ids = set(b_with_id["link_id"].unique())
    m_ids = set(m_with_id["link_id"].unique())
    result.broker_ids_count = len(b_ids)
    result.murex_ids_count = len(m_ids)

    common_ids = b_ids & m_ids
    only_broker = b_ids - m_ids
    only_murex = m_ids - b_ids

    # --- Step 1: compare every ID present on both sides ----------------------
    for link_id in common_ids:
        b_rows = b_with_id[b_with_id["link_id"] == link_id]
        m_rows = m_with_id[m_with_id["link_id"] == link_id]
        outcome = _compare_id_group(link_id, b_rows, m_rows, config)
        if outcome["status"] == "PASS":
            if outcome["record"]["is_aggregated"]:
                result.id_passed_aggregated.append(outcome["record"])
            else:
                result.id_passed.append(outcome["record"])
        else:
            result.id_failed.append(outcome["record"])

    # --- Step 2: IDs only on one side => missing ------------------------------
    for link_id in only_broker:
        b_rows = b_with_id[b_with_id["link_id"] == link_id]
        result.missing_in_murex.append({
            "link_id": link_id,
            "broker_legs": _fmt_many(b_rows, b_rows.index),
            "broker_qty_total": b_rows["quantity"].sum(),
        })
    for link_id in only_murex:
        m_rows = m_with_id[m_with_id["link_id"] == link_id]
        result.missing_in_broker.append({
            "link_id": link_id,
            "murex_legs": _fmt_many(m_rows, m_rows.index),
            "murex_qty_total": m_rows["quantity"].sum(),
        })

    # --- Step 3: fallback attribute matching for rows with no usable ID ------
    _fallback_attribute_match(b_without_id, m_without_id, config, result)

    return result


# ==============================================================================
# 3B. CROSS-ID AGGREGATION  -  OPTIONAL ENHANCEMENT LAYER (Section 7)
# ==============================================================================
# Everything in this section is ADDITIVE and runs strictly AFTER reconcile()
# above has already produced its result. Nothing in this section modifies
# any existing function, and nothing here can ever change an existing PASS.
#
# Business rule (confirmed and signed off before this was written):
#   Trigger   : an ID-matched pair in result.id_failed whose issue list
#               includes a "quantity" mismatch (i.e. the Murex side and the
#               broker side both exist under the SAME link_id, but the lots
#               don't add up).
#   Filters   : among broker rows NOT already consumed by any existing
#               PASS / aggregated match, filter by (1) trader - exact match,
#               required; (2) price - exact match, no tolerance; (3) if step
#               1+2 still leaves more than one candidate row, additionally
#               filter by commodity/instrument.
#   Success   : exactly ONE combination (subset) of the filtered candidate
#               rows sums exactly to the Murex quantity for that ID.
#   Tie / no  : if zero combinations or MORE THAN ONE combination sum
#   match     : exactly to the target, the case is left completely
#               untouched as a FAIL for manual review - this layer never
#               guesses between multiple equally-valid combinations.
#   Trader    : if the trader name is missing on the failed case itself,
#   missing   : the case is left untouched and a comment is attached
#               noting the trader name was missing, rather than silently
#               skipping it.
#   Output    : resolved cases are REMOVED from result.id_failed and from
#               the corresponding result.missing_in_murex entries, and
#               added to result.cross_id_resolved, tagged so they remain
#               distinguishable from a direct ID-based PASS.

def _find_exact_subset_unique(
    items: List[Tuple[str, float]], target: float, tol: float,
    limit: int, scale: int = 100
) -> Tuple[bool, List[str]]:
    """
    Determine whether there is EXACTLY ONE subset of `items` whose
    quantities sum to `target` (within `tol`), and if so return it.

    Returns (True, winning_keys) when exactly one subset exists.
    Returns (False, []) when zero OR more than one subset exists.
    The caller leaves ambiguous/no-match cases as FAIL (Q7 sign-off).

    ALGORITHM — two-pass 0/1 knapsack, O(n * cap) time, O(cap) space:

    Pass 1 — COUNT only (no combination storage):
      dp_count[s] = number of distinct subsets summing to s, capped at 2.
      Backwards sweep per item guarantees each item used at most once and
      each subset counted exactly once.  If the target count is not 1,
      return immediately — no reconstruction needed.

    Pass 2 — RECONSTRUCT (only when count == 1):
      Re-run the knapsack item-by-item, keeping a single 1-D boolean
      array. Before processing item i, snapshot the array. After
      processing, compare: if dp[winning_sum] was False before and True
      after, item i was taken. Walk backwards to recover every taken item.
      Space is O(cap) — no 2-D table, no exponential blowup.

    `limit` caps items considered as a safety ceiling (not a business rule).
    """
    if len(items) > limit:
        items = items[:limit]
    if not items:
        return False, []

    target_i = _scale_int(target, scale)
    tol_i    = max(0, _scale_int(tol, scale))
    cap      = target_i + tol_i

    # ------------------------------------------------------------------
    # Pass 1: count subsets (capped at 2) using backwards 0/1 knapsack
    # ------------------------------------------------------------------
    dp_count = [0] * (cap + 1)
    dp_count[0] = 1
    for _key, qty in items:
        q = _scale_int(qty, scale)
        if q <= 0:
            continue
        for s in range(cap, q - 1, -1):          # backwards → 0/1 knapsack
            if dp_count[s - q]:
                dp_count[s] = min(2, dp_count[s] + dp_count[s - q])

    total = sum(
        dp_count[s]
        for s in range(max(0, target_i - tol_i), cap + 1)
        if abs(s - target_i) <= tol_i
    )
    if total != 1:
        return False, []

    # ------------------------------------------------------------------
    # Pass 2: reconstruct the unique winning subset, O(cap) space
    # ------------------------------------------------------------------
    # dp_bool is a 1-D reachability array updated in-place per item.
    # To find the winning sum once, run the full knapsack first.
    dp_bool = [False] * (cap + 1)
    dp_bool[0] = True
    for _key, qty in items:
        q = _scale_int(qty, scale)
        if q <= 0:
            continue
        for s in range(cap, q - 1, -1):
            if dp_bool[s - q]:
                dp_bool[s] = True

    winning_sum = next(
        (s for s in range(max(0, target_i - tol_i), cap + 1)
         if abs(s - target_i) <= tol_i and dp_bool[s]),
        None,
    )
    if winning_sum is None:
        return False, []           # should not happen — pass 1 confirmed it

    # Recover which items were taken by replaying item-by-item with
    # a snapshot trick: if dp_pre[winning_sum - q] was True before
    # processing item i, then item i is needed to reach winning_sum.
    # We replay in reverse item order (last item first) to stay O(cap).
    winning_keys: List[str] = []
    remaining = winning_sum

    # Rebuild a reachability array EXCLUDING each item by replaying
    # all OTHER items — but that would be O(n²).  Instead use the
    # "before/after snapshot" approach in ONE forward pass:
    # keep a list of per-item snapshots of dp_bool BEFORE that item
    # was added, then backtrack.  Each snapshot is O(cap), total O(n*cap)
    # memory — but n here is the number of CANDIDATE rows (typically
    # small after trader+price filtering, rarely > 100 even for large files).
    snapshots: List[List[bool]] = []
    dp_snap = [False] * (cap + 1)
    dp_snap[0] = True
    for _key, qty in items:
        q = _scale_int(qty, scale)
        if q <= 0:
            snapshots.append(None)
            continue
        snapshots.append(dp_snap[:])         # snapshot BEFORE this item
        for s in range(cap, q - 1, -1):
            if dp_snap[s - q]:
                dp_snap[s] = True

    # Backtrack using snapshots
    for i in range(len(items) - 1, -1, -1):
        row_key, qty = items[i]
        q = _scale_int(qty, scale)
        if q <= 0 or snapshots[i] is None:
            continue
        pre = snapshots[i]
        # Item i was taken iff:
        #   (a) remaining >= q, AND
        #   (b) remaining - q was reachable BEFORE this item was added
        if remaining >= q and pre[remaining - q]:
            winning_keys.append(row_key)
            remaining -= q
        if remaining == 0:
            break

    return True, winning_keys


def _resolve_cross_id_aggregation(broker_df: pd.DataFrame, murex_df: pd.DataFrame,
                                    result: ReconResult, config: ReconConfig) -> None:
    """Mutates `result` in place: moves cleanly-resolved cases out of
    id_failed / missing_in_murex and into cross_id_resolved. Leaves
    everything else untouched. See module-level comment block above for
    the full business rule this implements."""

    if not config.enable_cross_id_aggregation:
        return

    # Row lookup by _row_ref so we can go from a formatted leg dict (which
    # only carries _row_ref, not the original dataframe index) back to the
    # live broker_df row.
    b_lookup = {broker_df.at[i, "_row_ref"]: i for i in broker_df.index}

    # Track which broker rows are already "spoken for" by an existing
    # PASS / aggregated match or by an earlier cross-ID resolution in this
    # same run, so the same broker row is never reused twice.
    consumed_row_refs = set()
    for rec in result.id_passed + result.id_passed_aggregated:
        for leg in rec["broker_legs"]:
            consumed_row_refs.add(leg["_row_ref"])
    for rec in result.fallback_matched:
        consumed_row_refs.add(rec["broker"]["_row_ref"])
    for rec in result.fallback_aggregated:
        for leg in rec["broker_trades"]:
            consumed_row_refs.add(leg["_row_ref"])

    # Quick lookup: link_id -> missing_in_murex entry, so a resolved FAIL
    # case can also pull in (and remove) the orphaned broker row(s) that
    # were separately sitting in Missing in Murex under their own IDs.
    missing_by_id = {e["link_id"]: e for e in result.missing_in_murex}

    still_failed = []
    resolved_link_ids = []

    for fail_rec in result.id_failed:
        qty_issue = next((i for i in fail_rec.get("issues", []) if i["field"] == "quantity"), None)
        if qty_issue is None:
            still_failed.append(fail_rec)
            continue  # not a quantity-mismatch FAIL - out of scope for this layer

        target_qty = fail_rec["murex_qty_total"]
        trader = None
        existing_broker_row_refs = {leg["_row_ref"] for leg in fail_rec["broker_legs"]}
        for ref in existing_broker_row_refs:
            idx = b_lookup.get(ref)
            if idx is not None:
                raw_trader = broker_df.at[idx, "trader"]
                if raw_trader is not None and not (isinstance(raw_trader, float) and pd.isna(raw_trader)):
                    trader = str(raw_trader).strip() or None
                if trader:
                    break

        if not trader:
            fail_rec["cross_id_comment"] = (
                "Cross-ID aggregation could not be attempted: trader name is missing "
                "on the broker side of this failed case."
            )
            still_failed.append(fail_rec)
            continue

        price = fail_rec.get("broker_price")
        commodity = fail_rec["broker_legs"][0].get("instrument") if fail_rec["broker_legs"] else None

        # Candidate pool: orphaned broker rows (not already consumed, not
        # already part of this failing ID itself) matching trader + price
        # exactly. Date/direction are NOT used as hard filters per business
        # sign-off, but a mismatch on either is recorded as a warning note
        # on the resolved record rather than excluding the row.
        candidates = broker_df[
            (broker_df["trader"] == trader)
            & (broker_df["price"].notna())
            & (broker_df["price"].sub(price).abs() <= CROSS_ID_PRICE_TOLERANCE)
            & (~broker_df["_row_ref"].isin(consumed_row_refs))
            & (~broker_df["_row_ref"].isin(existing_broker_row_refs))
        ]

        # If trader + price alone leaves more than one DISTINCT candidate
        # row, narrow further by commodity/instrument, per business rule
        # ("first trader, then price, and last by commodity if needed").
        if len(candidates) > 1 and commodity:
            narrowed = candidates[candidates["instrument"] == commodity]
            if len(narrowed) > 0:
                candidates = narrowed

        if candidates.empty:
            still_failed.append(fail_rec)
            continue

        # IMPORTANT: the failing broker leg(s) under this ID are already
        # part of the total (e.g. 25 of the 200 lots are already accounted
        # for by the ID match itself). The search below only needs to find
        # OTHER orphaned rows that make up the REMAINDER, not the full
        # Murex quantity again.
        already_have = fail_rec["broker_qty_total"]
        remaining_target = target_qty - already_have

        if remaining_target <= 0:
            # Broker side already has as much or more than Murex - this is
            # a genuine over-count, not something this layer should touch.
            still_failed.append(fail_rec)
            continue

        items = [(str(r), candidates.at[r, "quantity"]) for r in candidates.index]
        is_unique, winning_keys = _find_exact_subset_unique(
            items, remaining_target, config.quantity_tolerance,
            limit=config.cross_id_subset_sum_item_limit,
        )

        if not is_unique:
            # Zero matches OR a genuine tie (multiple valid combinations).
            # Either way: leave as FAIL for manual review - never guess (Q7).
            # Attach an explanatory comment so the analyst knows why.
            total_candidates = len(items)
            fail_rec["cross_id_comment"] = (
                f"Cross-ID aggregation searched {total_candidates} candidate broker row(s) "
                f"(trader='{trader}', price={price}) for a unique subset summing to "
                f"{remaining_target} lots (the {target_qty} Murex total minus the "
                f"{already_have} already matched). Either no combination found or more "
                "than one valid combination exists - left as FAIL for manual review."
            )
            still_failed.append(fail_rec)
            continue

        # Exactly one unambiguous combination found - resolve it.
        chosen_idx = [int(r) for r in winning_keys]
        all_broker_idx_for_record = [b_lookup[ref] for ref in existing_broker_row_refs if ref in b_lookup] + chosen_idx

        warning_notes = []
        b_dates = broker_df.loc[all_broker_idx_for_record, "trade_date"].dropna().unique()
        b_directions = broker_df.loc[all_broker_idx_for_record, "direction"].dropna().unique()
        if len(b_dates) > 1:
            warning_notes.append("Trade dates differ across the combined broker rows.")
        if len(b_directions) > 1:
            warning_notes.append("Direction (buy/sell) differs across the combined broker rows.")
        m_dates = pd.Series([leg.get("trade_date") for leg in fail_rec["murex_legs"]]).dropna().unique()
        if len(b_dates) == 1 and len(m_dates) == 1 and str(b_dates[0])[:10] != str(m_dates[0])[:10]:
            warning_notes.append("Trade date differs from the Murex side.")

        resolved_record = {
            "link_id": fail_rec["link_id"],
            "resolution_type": "Resolved by Cross-ID Aggregation",
            "matched_by": {"trader": trader, "price": price, "commodity": commodity},
            "broker_legs": _fmt_many(broker_df, all_broker_idx_for_record),
            "murex_legs": fail_rec["murex_legs"],
            "broker_qty_total": broker_df.loc[all_broker_idx_for_record, "quantity"].sum(),
            "murex_qty_total": target_qty,
            "warning_notes": warning_notes,
        }
        result.cross_id_resolved.append(resolved_record)
        resolved_link_ids.append(fail_rec["link_id"])
        consumed_row_refs.update(existing_broker_row_refs)
        consumed_row_refs.update(broker_df.at[i, "_row_ref"] for i in chosen_idx)

        # Remove any Missing-in-Murex entries for the orphaned rows we just
        # consumed, since they are now accounted for by this resolution.
        chosen_row_refs = {broker_df.at[i, "_row_ref"] for i in chosen_idx}
        result.missing_in_murex = [
            e for e in result.missing_in_murex
            if not set(leg["_row_ref"] for leg in e["broker_legs"]).issubset(chosen_row_refs)
        ]

    result.id_failed = still_failed


# ==============================================================================
# 3C. MISSING-SET ATTRIBUTE AGGREGATION  -  NEW ENHANCEMENT LAYER (Section 8)
# ==============================================================================
# Runs strictly AFTER reconcile() AND after _resolve_cross_id_aggregation().
# It never touches an existing PASS or FAIL entry.
#
# Problem it solves
# -----------------
# In some real-world file pairs (e.g. Tullett Prebon vs Murex) the Murex G.ID
# is a *parent* identifier that has NO exact match among the broker's granular
# per-execution IDs (e.g. Murex 'ID17818888630000' vs broker rows like
# 'ID178188886300000001', 'ID178188886300000002', ...).  Because the IDs never
# intersect, reconcile() dumps ALL broker rows into Missing-in-Murex and ALL
# Murex rows into Missing-in-Broker.  The cross-ID layer in 3B never fires
# because there are no id_failed records.
#
# Business rule (documented & agreed):
#   For each Murex entry sitting in Missing-in-Broker:
#     1. Extract its price, trader, commodity and target qty.
#     2. Among broker rows currently in Missing-in-Murex (and not yet consumed),
#        filter by: exact price first, then trader (exact), then commodity if
#        still > 1 candidate.
#     3. Find whether exactly ONE subset of the filtered candidates sums to the
#        Murex qty (same knapsack uniqueness algorithm as cross-ID).
#     4. If EXACTLY ONE combination exists => move both sides to
#        result.missing_attr_resolved, remove from both missing lists.
#     5. Zero matches OR tie => leave as Missing with an explanatory comment.
# ==============================================================================

def _resolve_missing_by_attribute_aggregation(
        broker_df: pd.DataFrame, murex_df: pd.DataFrame,
        result: ReconResult, config: ReconConfig) -> None:
    """Mutates `result` in-place: moves cleanly-resolved missing entries out of
    missing_in_murex / missing_in_broker and into missing_attr_resolved.
    Everything else is left completely untouched."""

    if not config.enable_missing_attr_resolution:
        return

    # Build _row_ref -> dataframe index lookup for fast attribute access.
    b_lookup = {broker_df.at[i, "_row_ref"]: i for i in broker_df.index}

    # Which broker _row_refs are currently orphaned (in missing_in_murex)?
    orphan_broker_refs = set()
    for entry in result.missing_in_murex:
        for leg in entry["broker_legs"]:
            orphan_broker_refs.add(leg["_row_ref"])

    # Broker row refs consumed by this layer (so the same row is never reused).
    consumed_broker_refs = set()

    still_missing_broker = []

    for murex_entry in result.missing_in_broker:
        m_link_id = murex_entry.get("link_id")
        unresolved_legs = []
        unresolved_comments = []

        for m_leg in murex_entry["murex_legs"]:
            target_qty = m_leg.get("quantity")
            try:
                target_qty = float(target_qty)
            except (TypeError, ValueError):
                target_qty = None

            price = m_leg.get("price")
            try:
                price = float(price)
            except (TypeError, ValueError):
                price = None

            raw_trader = m_leg.get("trader")
            trader = str(raw_trader).strip().upper() if raw_trader else None
            trader = trader or None

            raw_commodity = m_leg.get("instrument")
            commodity = str(raw_commodity).strip().upper() if raw_commodity else None
            commodity = commodity or None

            # Price and quantity are minimum required fields for safe matching.
            if price is None or target_qty is None or target_qty <= 0:
                unresolved_legs.append(m_leg)
                unresolved_comments.append(
                    f"Leg row {m_leg.get('_row_ref')}: missing/invalid price or quantity."
                )
                continue

            # --- Candidate pool: orphaned broker rows not yet consumed -------
            available_refs = orphan_broker_refs - consumed_broker_refs
            candidate_idx = [b_lookup[ref] for ref in available_refs if ref in b_lookup]
            if not candidate_idx:
                unresolved_legs.append(m_leg)
                unresolved_comments.append(
                    f"Leg row {m_leg.get('_row_ref')}: no broker orphan rows available."
                )
                continue

            candidates = broker_df.loc[candidate_idx].copy()

            # Optional narrowing: prefer broker child IDs under this Murex parent ID.
            if m_link_id is not None:
                m_parent = str(m_link_id).strip().upper()
                if m_parent:
                    family = candidates[
                        candidates["link_id"].astype(str).str.strip().str.upper().str.startswith(m_parent)
                    ]
                    if not family.empty:
                        candidates = family

            # Filter 1: exact price
            candidates = candidates[
                candidates["price"].notna() &
                (candidates["price"].sub(price).abs() <= CROSS_ID_PRICE_TOLERANCE)
            ]
            if candidates.empty:
                unresolved_legs.append(m_leg)
                unresolved_comments.append(
                    f"Leg row {m_leg.get('_row_ref')}: no broker candidates at price {price}."
                )
                continue

            # Filter 2: trader (if available on Murex leg)
            if trader:
                narrowed = candidates[candidates["trader"] == trader]
                if not narrowed.empty:
                    candidates = narrowed

            # Filter 3: commodity (only if still ambiguous)
            if len(candidates) > 1 and commodity:
                narrowed = candidates[candidates["instrument"] == commodity]
                if not narrowed.empty:
                    candidates = narrowed

            # --- Unique subset-sum check for THIS leg -----------------------
            items = [(str(r), candidates.at[r, "quantity"]) for r in candidates.index]
            is_unique, winning_keys = _find_exact_subset_unique(
                items, target_qty, config.quantity_tolerance,
                limit=config.cross_id_subset_sum_item_limit,
            )

            if not is_unique:
                unresolved_legs.append(m_leg)
                unresolved_comments.append(
                    f"Leg row {m_leg.get('_row_ref')}: searched {len(items)} candidate row(s) "
                    f"(price={price}, trader='{trader}', commodity='{commodity}') for unique "
                    f"subset = {target_qty}; no unique solution."
                )
                continue

            # --- Resolved leg -----------------------------------------------
            chosen_idx = [int(r) for r in winning_keys]
            chosen_refs = {broker_df.at[i, "_row_ref"] for i in chosen_idx}
            consumed_broker_refs.update(chosen_refs)

            warning_notes = []
            b_dates = broker_df.loc[chosen_idx, "trade_date"].dropna().unique()
            b_directions = broker_df.loc[chosen_idx, "direction"].dropna().unique()
            if len(b_dates) > 1:
                warning_notes.append("Trade dates differ across the combined broker rows.")
            if len(b_directions) > 1:
                warning_notes.append("Direction (buy/sell) differs across the combined broker rows.")
            m_leg_date = m_leg.get("trade_date")
            if len(b_dates) == 1 and m_leg_date:
                if str(b_dates[0])[:10] != str(m_leg_date)[:10]:
                    warning_notes.append("Trade date differs between broker and Murex side.")

            resolved_record = {
                "murex_link_id": m_link_id,
                "resolution_type": "Resolved by Attribute Aggregation (Missing sets)",
                "matched_by": {"trader": trader, "price": price, "commodity": commodity},
                "broker_legs": _fmt_many(broker_df, chosen_idx),
                "murex_legs": [m_leg],
                "broker_qty_total": float(broker_df.loc[chosen_idx, "quantity"].sum()),
                "murex_qty_total": target_qty,
                "warning_notes": warning_notes,
            }
            result.missing_attr_resolved.append(resolved_record)

            # Remove consumed broker rows from missing_in_murex.
            result.missing_in_murex = [
                e for e in result.missing_in_murex
                if not set(leg["_row_ref"] for leg in e["broker_legs"]).issubset(chosen_refs)
            ]
            orphan_broker_refs -= chosen_refs

        # Keep only unresolved legs (if any) in Missing in Broker.
        if unresolved_legs:
            still_missing_broker.append({
                "link_id": m_link_id,
                "murex_legs": unresolved_legs,
                "murex_qty_total": float(sum((leg.get("quantity") or 0.0) for leg in unresolved_legs)),
                "attr_match_comment": " | ".join(unresolved_comments[:5]),
            })

    result.missing_in_broker = still_missing_broker


# ==============================================================================
# 4. EXCEL REPORT BUILDER
# ==============================================================================

FONT_NAME = "Arial"
HEADER_FILL = PatternFill("solid", start_color="1F4E78")
HEADER_FONT = Font(name=FONT_NAME, bold=True, color="FFFFFF")
TITLE_FONT = Font(name=FONT_NAME, bold=True, size=14)
SUBTITLE_FONT = Font(name=FONT_NAME, bold=True, size=11)
BOLD = Font(name=FONT_NAME, bold=True)
NORMAL = Font(name=FONT_NAME)
PASS_FILL = PatternFill("solid", start_color="C6EFCE")
FAIL_FILL = PatternFill("solid", start_color="FFC7CE")
WARN_FILL = PatternFill("solid", start_color="FFEB9C")
THIN = Side(border_style="thin", color="B0B0B0")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
CENTER = Alignment(horizontal="center", vertical="center")

LEG_COLS = ["trade_id", "link_id", "instrument", "direction", "trade_date",
            "prompt_date", "price", "quantity", "currency", "broker",
            "counterparty", "trader", "_row_ref"]
LEG_HEADERS = ["Trade ID", "Link/Trade ID", "Instrument", "Direction", "Trade Date",
               "Prompt Date", "Price", "Quantity/Lots", "Currency", "Broker",
               "Counterparty", "Trader", "Source Row"]


def _legs_to_text(legs: List[dict], field_name: str) -> str:
    return ", ".join(str(l.get(field_name)) for l in legs)


def _write_header(ws, headers, row=1):
    for c, h in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=c, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        cell.border = BORDER


def _autosize(ws, max_width=45):
    for col_cells in ws.columns:
        length = max((len(str(c.value)) for c in col_cells if c.value is not None), default=8)
        ws.column_dimensions[get_column_letter(col_cells[0].column)].width = min(max_width, max(10, length + 2))


def _apply_borders(ws, n_rows, n_cols, start_row=2):
    for r in range(start_row, start_row + n_rows):
        for c in range(1, n_cols + 1):
            ws.cell(row=r, column=c).font = NORMAL
            ws.cell(row=r, column=c).border = BORDER


def _build_summary(wb, result: ReconResult, config: ReconConfig):
    ws = wb.active
    ws.title = "Executive Summary"
    ws["A1"] = "Trade Reconciliation - Executive Summary"
    ws["A1"].font = TITLE_FONT
    ws.merge_cells("A1:C1")

    n_pass = len(result.id_passed) + len(result.id_passed_aggregated)
    n_fail = len(result.id_failed)
    n_missing_murex = len(result.missing_in_murex)
    n_missing_broker = len(result.missing_in_broker)
    n_cross_id_resolved = len(result.cross_id_resolved)
    n_missing_attr_resolved = len(result.missing_attr_resolved)
    # Each resolved CASE can cover multiple broker IDs (e.g. one Murex ID
    # resolved by combining 14 different broker IDs) - for the "% of broker
    # IDs reconciled" status line, count broker IDs covered, not cases.
    n_cross_id_resolved_broker_ids = sum(len(r["broker_legs"]) for r in result.cross_id_resolved)
    n_missing_attr_resolved_broker_ids = sum(len(r["broker_legs"]) for r in result.missing_attr_resolved)

    rows = [
        ("Total trade rows in Broker File", result.broker_count),
        ("Total trade rows in Murex File", result.murex_count),
        ("Unique Trade/Link IDs in Broker File", result.broker_ids_count),
        ("Unique Trade/Link IDs in Murex File", result.murex_ids_count),
        ("", ""),
        ("IDs Reconciled - PASS (single leg)", len(result.id_passed)),
        ("IDs Reconciled - PASS (aggregated / multi-leg)", len(result.id_passed_aggregated)),
        ("IDs Resolved - PASS (Cross-ID Aggregation, enhancement layer)", n_cross_id_resolved),
        ("IDs Resolved - PASS (Attribute Aggregation, missing-set layer)", n_missing_attr_resolved),
        ("IDs Reconciled - FAIL (field mismatch)", n_fail),
        ("IDs Missing in Murex (present in Broker only)", n_missing_murex),
        ("IDs Missing in Broker (present in Murex only)", n_missing_broker),
        ("", ""),
        ("Rows with no usable Trade/Link ID - Broker", len(result.no_id_unmatched_broker) +
         len(result.fallback_matched) + sum(len(a["broker_trades"]) for a in result.fallback_aggregated)),
        ("  - Fallback matched (1:1)", len(result.fallback_matched)),
        ("  - Fallback aggregated", len(result.fallback_aggregated)),
        ("  - Fallback unmatched (true exceptions)", len(result.no_id_unmatched_broker)),
    ]

    r = 3
    ws.cell(row=r, column=1, value="Metric").font = HEADER_FONT
    ws.cell(row=r, column=1).fill = HEADER_FILL
    ws.cell(row=r, column=2, value="Value").font = HEADER_FONT
    ws.cell(row=r, column=2).fill = HEADER_FILL
    r += 1
    for label, value in rows:
        if label == "":
            r += 1
            continue
        ws.cell(row=r, column=1, value=label).font = BOLD if not label.startswith("  -") else NORMAL
        c = ws.cell(row=r, column=2, value=value)
        c.alignment = CENTER
        r += 1

    total_ids = result.broker_ids_count
    r += 1
    ws.cell(row=r, column=1, value="Overall Status").font = SUBTITLE_FONT
    r += 1
    # n_cross_id_resolved is always 0 when the enhancement layer is disabled
    # or finds nothing, so this calculation is identical to the original
    # behaviour in that case.
    pct_clean = 100.0 * (n_pass + n_cross_id_resolved_broker_ids + n_missing_attr_resolved_broker_ids) / total_ids if total_ids else 0.0
    status_cell = ws.cell(row=r, column=1, value=f"{pct_clean:.1f}% of Broker trade IDs reconciled cleanly (PASS)")
    status_cell.fill = PASS_FILL if pct_clean >= 95 else (WARN_FILL if pct_clean >= 80 else FAIL_FILL)
    status_cell.font = BOLD
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)

    r += 2
    ws.cell(row=r, column=1, value="Matching Configuration").font = SUBTITLE_FONT
    r += 1
    ws.cell(row=r, column=1, value="Primary match key").font = BOLD
    ws.cell(row=r, column=2, value="Broker link_id <-> Murex link_id (e.g. Unique Trade ID <-> G.ID)")
    r += 1
    ws.cell(row=r, column=1, value="Price Tolerance").font = BOLD
    ws.cell(row=r, column=2, value=config.price_tolerance)
    r += 1
    ws.cell(row=r, column=1, value="Quantity Tolerance").font = BOLD
    ws.cell(row=r, column=2, value=config.quantity_tolerance)
    r += 1
    ws.cell(row=r, column=1, value="Fallback attribute matching (no-ID rows)").font = BOLD
    ws.cell(row=r, column=2, value=str(config.enable_fallback_matching))
    r += 1
    ws.cell(row=r, column=1, value="Fallback group/match keys").font = BOLD
    ws.cell(row=r, column=2, value=", ".join(config.group_keys))
    r += 1
    ws.cell(row=r, column=1, value="Cross-ID aggregation (enhancement layer)").font = BOLD
    ws.cell(row=r, column=2, value=str(config.enable_cross_id_aggregation))
    r += 1
    ws.cell(row=r, column=1, value="Attribute aggregation for missing sets (Section 8)").font = BOLD
    ws.cell(row=r, column=2, value=str(config.enable_missing_attr_resolution))

    ws.column_dimensions["A"].width = 48
    ws.column_dimensions["B"].width = 32
    ws.column_dimensions["C"].width = 20


def _build_id_recon(wb, result: ReconResult):
    """Sheet 2: ID-Level Reconciliation - one row per link_id, PASS or FAIL,
    showing broker total vs murex total and any field-level issues."""
    ws = wb.create_sheet("ID-Level Reconciliation")
    headers = ["Link/Trade ID", "Status", "Leg Type", "Broker Leg Count", "Broker Direction",
               "Broker Price", "Broker Qty Total", "Murex Leg Count", "Murex Direction",
               "Murex Price", "Murex Qty Total", "Issue(s)"]
    _write_header(ws, headers)
    r = 2

    def add_row(rec, status):
        leg_type = "Aggregated (multi-leg)" if rec["is_aggregated"] else "Single leg"
        issue_text = ""
        if status == "FAIL":
            issue_text = "; ".join(f"{i['reason']} (Broker: {i['broker_value']} | Murex: {i['murex_value']})"
                                    for i in rec["issues"])
        ws.append([
            rec["link_id"], status, leg_type, len(rec["broker_legs"]), rec["broker_direction"],
            rec["broker_price"], rec["broker_qty_total"], len(rec["murex_legs"]), rec["murex_direction"],
            rec["murex_price"], rec["murex_qty_total"], issue_text,
        ])
        nonlocal r
        fill = PASS_FILL if status == "PASS" else FAIL_FILL
        for c in range(1, len(headers) + 1):
            ws.cell(row=r, column=c).fill = fill
        r += 1

    for rec in result.id_passed:
        add_row(rec, "PASS")
    for rec in result.id_passed_aggregated:
        add_row(rec, "PASS")
    for rec in result.id_failed:
        add_row(rec, "FAIL")

    _apply_borders(ws, len(result.id_passed) + len(result.id_passed_aggregated) + len(result.id_failed), len(headers))
    _autosize(ws)
    ws.freeze_panes = "A2"


def _build_passed(wb, result: ReconResult):
    ws = wb.create_sheet("Matched (PASS)")
    headers = (["Link/Trade ID", "Leg Type"]
               + [f"Broker {h}" for h in LEG_HEADERS]
               + [f"Murex {h}" for h in LEG_HEADERS])
    _write_header(ws, headers)
    r = 2
    all_pass = result.id_passed + result.id_passed_aggregated
    for rec in all_pass:
        b_legs, m_legs = rec["broker_legs"], rec["murex_legs"]
        n_lines = max(len(b_legs), len(m_legs))
        for i in range(n_lines):
            b = b_legs[i] if i < len(b_legs) else {c: "" for c in LEG_COLS}
            m = m_legs[i] if i < len(m_legs) else {c: "" for c in LEG_COLS}
            leg_type = "Aggregated" if rec["is_aggregated"] else "Single leg"
            ws.append([rec["link_id"] if i == 0 else "", leg_type if i == 0 else ""]
                      + [b.get(c) for c in LEG_COLS] + [m.get(c) for c in LEG_COLS])
            for c in range(1, len(headers) + 1):
                ws.cell(row=r, column=c).fill = PASS_FILL
            r += 1
    _apply_borders(ws, r - 2, len(headers))
    _autosize(ws)
    ws.freeze_panes = "A2"


def _build_failed(wb, result: ReconResult):
    ws = wb.create_sheet("Mismatched (FAIL)")
    headers = (["Link/Trade ID", "Reason(s)", "Broker Value(s)", "Murex Value(s)"]
               + [f"Broker {h}" for h in LEG_HEADERS]
               + [f"Murex {h}" for h in LEG_HEADERS])
    _write_header(ws, headers)
    r = 2
    for rec in result.id_failed:
        reasons = "; ".join(i["reason"] for i in rec["issues"])
        bvals = "; ".join(str(i["broker_value"]) for i in rec["issues"])
        mvals = "; ".join(str(i["murex_value"]) for i in rec["issues"])
        b_legs, m_legs = rec["broker_legs"], rec["murex_legs"]
        n_lines = max(len(b_legs), len(m_legs))
        for i in range(n_lines):
            b = b_legs[i] if i < len(b_legs) else {c: "" for c in LEG_COLS}
            m = m_legs[i] if i < len(m_legs) else {c: "" for c in LEG_COLS}
            ws.append([rec["link_id"] if i == 0 else "", reasons if i == 0 else "",
                      bvals if i == 0 else "", mvals if i == 0 else ""]
                      + [b.get(c) for c in LEG_COLS] + [m.get(c) for c in LEG_COLS])
            for c in range(1, len(headers) + 1):
                ws.cell(row=r, column=c).fill = FAIL_FILL
            r += 1
    _apply_borders(ws, r - 2, len(headers))
    _autosize(ws)
    ws.freeze_panes = "A2"


def _build_missing(wb, title, entries, side_key, qty_key):
    ws = wb.create_sheet(title)
    headers = ["Link/Trade ID", "Leg Count", "Qty Total"] + LEG_HEADERS
    _write_header(ws, headers)
    r = 2
    for e in entries:
        legs = e[side_key]
        for i, leg in enumerate(legs):
            ws.append([e["link_id"] if i == 0 else "", len(legs) if i == 0 else "",
                      e[qty_key] if i == 0 else ""] + [leg.get(c) for c in LEG_COLS])
            for c in range(1, len(headers) + 1):
                ws.cell(row=r, column=c).fill = WARN_FILL
            r += 1
    _apply_borders(ws, r - 2, len(headers))
    _autosize(ws)
    ws.freeze_panes = "A2"


def _build_aggregated_detail(wb, result: ReconResult):
    """Sheet showing exactly how broker legs rolled up per aggregated ID,
    fulfilling the 'aggregated match report' requirement explicitly."""
    ws = wb.create_sheet("Aggregated Matches Detail")
    headers = ["Link/Trade ID", "Broker Leg Count", "Broker Lots (each leg)",
               "Broker Lots Sum", "Murex Leg Count", "Murex Lots (each leg)",
               "Murex Lots Sum", "Difference", "Status"]
    _write_header(ws, headers)
    r = 2
    combined = [(rec, "PASS") for rec in result.id_passed_aggregated] + \
               [(rec, "FAIL") for rec in result.id_failed if rec["is_aggregated"]]
    for rec, status in combined:
        b_lots = ", ".join(str(l["quantity"]) for l in rec["broker_legs"])
        m_lots = ", ".join(str(l["quantity"]) for l in rec["murex_legs"])
        diff = round(rec["broker_qty_total"] - rec["murex_qty_total"], 4)
        ws.append([rec["link_id"], len(rec["broker_legs"]), b_lots, rec["broker_qty_total"],
                   len(rec["murex_legs"]), m_lots, rec["murex_qty_total"], diff, status])
        fill = PASS_FILL if status == "PASS" else FAIL_FILL
        for c in range(1, len(headers) + 1):
            ws.cell(row=r, column=c).fill = fill
        r += 1
    _apply_borders(ws, len(combined), len(headers))
    _autosize(ws)
    ws.freeze_panes = "A2"


def _build_fallback(wb, result: ReconResult):
    ws = wb.create_sheet("Exception Analysis (No-ID)")
    ws["A1"] = "Rows With No Usable Trade/Link ID"
    ws["A1"].font = TITLE_FONT
    ws.merge_cells("A1:F1")
    ws["A2"] = ("These rows had a blank or unmatched Trade/Link ID on at least one side, so "
                "ID-based reconciliation could not be applied. The fallback attribute matcher "
                "(instrument, direction, dates, price) was used instead - review these manually.")
    ws["A2"].font = Font(name=FONT_NAME, italic=True)
    ws.merge_cells("A2:F2")

    r = 4
    ws.cell(row=r, column=1, value="Fallback Matched (1:1, no ID)").font = SUBTITLE_FONT
    r += 1
    headers = [f"Broker {h}" for h in LEG_HEADERS] + [f"Murex {h}" for h in LEG_HEADERS]
    _write_header(ws, headers, row=r)
    r += 1
    start = r
    for m in result.fallback_matched:
        ws.append([m["broker"].get(c) for c in LEG_COLS] + [m["murex"].get(c) for c in LEG_COLS])
        for c in range(1, len(headers) + 1):
            ws.cell(row=r, column=c).fill = PASS_FILL
        r += 1
    _apply_borders(ws, r - start, len(headers), start_row=start)

    r += 2
    ws.cell(row=r, column=1, value="Fallback Aggregated (no ID)").font = SUBTITLE_FONT
    r += 1
    headers2 = ["Broker Trade Count", "Broker Qty Total", "Murex Trade Count", "Murex Qty Total"]
    _write_header(ws, headers2, row=r)
    r += 1
    start = r
    for a in result.fallback_aggregated:
        ws.append([len(a["broker_trades"]), a["broker_qty_total"], len(a["murex_trades"]), a["murex_qty_total"]])
        for c in range(1, len(headers2) + 1):
            ws.cell(row=r, column=c).fill = WARN_FILL
        r += 1
    _apply_borders(ws, r - start, len(headers2), start_row=start)

    r += 2
    ws.cell(row=r, column=1, value="Unmatched Broker rows (no ID, no fallback match found)").font = SUBTITLE_FONT
    r += 1
    _write_header(ws, LEG_HEADERS, row=r)
    r += 1
    start = r
    for leg in result.no_id_unmatched_broker:
        ws.append([leg.get(c) for c in LEG_COLS])
        for c in range(1, len(LEG_HEADERS) + 1):
            ws.cell(row=r, column=c).fill = FAIL_FILL
        r += 1
    _apply_borders(ws, r - start, len(LEG_HEADERS), start_row=start)

    r += 2
    ws.cell(row=r, column=1, value="Unmatched Murex rows (no ID, no fallback match found)").font = SUBTITLE_FONT
    r += 1
    _write_header(ws, LEG_HEADERS, row=r)
    r += 1
    start = r
    for leg in result.no_id_unmatched_murex:
        ws.append([leg.get(c) for c in LEG_COLS])
        for c in range(1, len(LEG_HEADERS) + 1):
            ws.cell(row=r, column=c).fill = FAIL_FILL
        r += 1
    _apply_borders(ws, r - start, len(LEG_HEADERS), start_row=start)

    _autosize(ws)


def _build_cross_id_resolved(wb, result: ReconResult):
    """Sheet 9 (additive): cases resolved by the Cross-ID Aggregation
    enhancement layer (Section 7). Kept as its own distinct sheet, never
    merged into the Matched (PASS) sheet, so a reviewer can always see
    which passes came from a direct ID match versus this attribute-based
    resolution."""
    ws = wb.create_sheet("Resolved - Cross-ID Aggregation")

    ws["A1"] = "Resolved by Cross-ID Aggregation"
    ws["A1"].font = TITLE_FONT
    ws.merge_cells("A1:F1")
    ws["A2"] = ("These cases originally failed on lots/quantity under their matched Link ID. "
                "Other broker rows under DIFFERENT IDs - filtered by matching trader and exact "
                "price - were found whose combined lots sum exactly to the Murex quantity, with "
                "no ambiguity (i.e. exactly one valid combination). They are treated as PASS and "
                "removed from the Mismatched and Missing in Murex sheets, but are kept on this "
                "separate sheet for full traceability.")
    ws["A2"].font = Font(name=FONT_NAME, italic=True)
    ws.merge_cells("A2:F2")

    headers = ["Link/Trade ID", "Matched By (Trader / Price / Commodity)",
               "Broker Leg Count", "Broker Lots (each leg)", "Broker Lots Sum",
               "Murex Lots Sum", "Warnings"]
    _write_header(ws, headers, row=4)
    r = 5
    for rec in result.cross_id_resolved:
        mb = rec["matched_by"]
        matched_by_text = f"{mb.get('trader')} / {mb.get('price')} / {mb.get('commodity') or '-'}"
        b_lots = ", ".join(str(l["quantity"]) for l in rec["broker_legs"])
        warnings_text = "; ".join(rec.get("warning_notes", [])) or ""
        ws.append([rec["link_id"], matched_by_text, len(rec["broker_legs"]), b_lots,
                   rec["broker_qty_total"], rec["murex_qty_total"], warnings_text])
        fill = WARN_FILL if warnings_text else PASS_FILL
        for c in range(1, len(headers) + 1):
            ws.cell(row=r, column=c).fill = fill
        r += 1
    _apply_borders(ws, len(result.cross_id_resolved), len(headers), start_row=5)
    _autosize(ws)
    ws.freeze_panes = "A5"

    # Full leg-level detail underneath, for anyone who needs to trace the
    # exact original rows that were combined.
    r += 2
    ws.cell(row=r, column=1, value="Leg-Level Detail").font = SUBTITLE_FONT
    r += 1
    detail_headers = [f"Broker {h}" for h in LEG_HEADERS] + [f"Murex {h}" for h in LEG_HEADERS]
    _write_header(ws, detail_headers, row=r)
    r += 1
    start = r
    for rec in result.cross_id_resolved:
        b_legs, m_legs = rec["broker_legs"], rec["murex_legs"]
        n_lines = max(len(b_legs), len(m_legs))
        for i in range(n_lines):
            b = b_legs[i] if i < len(b_legs) else {c: "" for c in LEG_COLS}
            m = m_legs[i] if i < len(m_legs) else {c: "" for c in LEG_COLS}
            ws.append([b.get(c) for c in LEG_COLS] + [m.get(c) for c in LEG_COLS])
            for c in range(1, len(detail_headers) + 1):
                ws.cell(row=r, column=c).fill = PASS_FILL
            r += 1
    _apply_borders(ws, r - start, len(detail_headers), start_row=start)
    _autosize(ws)


def _build_missing_attr_resolved(wb, result: ReconResult):
    """Sheet: cases resolved by the Missing-set Attribute Aggregation layer
    (Section 8). Kept separate from 'Matched (PASS)' for full traceability."""
    ws = wb.create_sheet("Resolved - Attr Aggregation")
    ws["A1"] = "Resolved by Attribute Aggregation (Missing Sets)"
    ws["A1"].font = TITLE_FONT
    ws.merge_cells("A1:F1")
    ws["A2"] = (
        "These Murex entries had no direct ID match in the broker file. "
        "Broker rows sitting in the Missing-in-Murex pool were filtered by matching price "
        "(exact), then trader, then commodity, and the unique combination whose lots sum "
        "equals the Murex quantity was accepted as a PASS. Multiple valid combinations "
        "are always left as Missing for manual review."
    )
    ws["A2"].font = Font(name=FONT_NAME, italic=True)
    ws.merge_cells("A2:F2")

    headers = ["Murex Link ID", "Matched By (Trader / Price / Commodity)",
               "Broker Leg Count", "Broker Lots (each leg)", "Broker Lots Sum",
               "Murex Lots Sum", "Warnings"]
    _write_header(ws, headers, row=4)
    r = 5
    for rec in result.missing_attr_resolved:
        mb = rec["matched_by"]
        matched_by_text = f"{mb.get('trader')} / {mb.get('price')} / {mb.get('commodity') or '-'}"
        b_lots = ", ".join(str(l["quantity"]) for l in rec["broker_legs"])
        warnings_text = "; ".join(rec.get("warning_notes", [])) or ""
        ws.append([rec["murex_link_id"], matched_by_text, len(rec["broker_legs"]),
                   b_lots, rec["broker_qty_total"], rec["murex_qty_total"], warnings_text])
        fill = WARN_FILL if warnings_text else PASS_FILL
        for c in range(1, len(headers) + 1):
            ws.cell(row=r, column=c).fill = fill
        r += 1
    _apply_borders(ws, len(result.missing_attr_resolved), len(headers), start_row=5)
    _autosize(ws)
    ws.freeze_panes = "A5"

    # Leg-level detail below the summary block
    r += 2
    ws.cell(row=r, column=1, value="Leg-Level Detail").font = SUBTITLE_FONT
    r += 1
    detail_headers = [f"Broker {h}" for h in LEG_HEADERS] + [f"Murex {h}" for h in LEG_HEADERS]
    _write_header(ws, detail_headers, row=r)
    r += 1
    start = r
    for rec in result.missing_attr_resolved:
        b_legs, m_legs = rec["broker_legs"], rec["murex_legs"]
        n_lines = max(len(b_legs), len(m_legs))
        for i in range(n_lines):
            b = b_legs[i] if i < len(b_legs) else {c: "" for c in LEG_COLS}
            m = m_legs[i] if i < len(m_legs) else {c: "" for c in LEG_COLS}
            ws.append([b.get(c) for c in LEG_COLS] + [m.get(c) for c in LEG_COLS])
            for c in range(1, len(detail_headers) + 1):
                ws.cell(row=r, column=c).fill = PASS_FILL
            r += 1
    _apply_borders(ws, r - start, len(detail_headers), start_row=start)
    _autosize(ws)


def build_report(result: ReconResult, config: ReconConfig, output_path: str):
    wb = Workbook()
    _build_summary(wb, result, config)
    _build_id_recon(wb, result)
    _build_passed(wb, result)
    _build_failed(wb, result)
    _build_missing(wb, "Missing in Murex", result.missing_in_murex, "broker_legs", "broker_qty_total")
    _build_missing(wb, "Missing in Broker", result.missing_in_broker, "murex_legs", "murex_qty_total")
    _build_aggregated_detail(wb, result)
    _build_fallback(wb, result)
    if result.cross_id_resolved:
        _build_cross_id_resolved(wb, result)
    if result.missing_attr_resolved:
        _build_missing_attr_resolved(wb, result)
    wb.save(output_path)
    return output_path


# ==============================================================================
# 5. STREAMLIT UI
# ==============================================================================

st.set_page_config(page_title="Trade Reconciliation Utility", layout="wide")
st.title("🔄 Trade Reconciliation Utility")
st.caption("ID-first reconciliation: Broker Unique Trade ID ↔ Murex G.ID (or equivalent), "
           "with lots aggregation, field-level mismatch detection, and an attribute-based fallback.")

col1, col2 = st.columns(2)
with col1:
    broker_file = st.file_uploader("📥 Broker File (SFTP source)", type=["csv", "xlsx", "xls"])
with col2:
    murex_file = st.file_uploader("📥 Murex Extract File", type=["csv", "xlsx", "xls"])

if broker_file and murex_file:
    broker_raw = read_any(broker_file)
    murex_raw = read_any(murex_file)

    st.success(f"Broker file: {len(broker_raw)} rows  |  Murex file: {len(murex_raw)} rows")

    with st.expander("Preview - Broker File", expanded=False):
        st.dataframe(broker_raw.head(20))
    with st.expander("Preview - Murex File", expanded=False):
        st.dataframe(murex_raw.head(20))

    st.subheader("⚙️ Column Mapping")
    st.caption("The **Link/Trade ID** mapping is the most important field — it is the primary "
               "key used to reconcile trades between the two files (e.g. Broker 'Unique Trade ID' "
               "/ 'Orig ID' ↔ Murex 'G.ID'). Set the other fields for accurate field-level checks.")

    map_col1, map_col2 = st.columns(2)
    broker_map, murex_map = {}, {}

    with map_col1:
        st.markdown("**Broker File columns**")
        for fname in CANONICAL_FIELDS:
            options = ["(None)"] + list(broker_raw.columns)
            default = DEFAULT_BROKER_COLUMN_MAP.get(fname)
            idx = options.index(default) if default in options else 0
            choice = st.selectbox(f"{fname}", options, index=idx, key=f"b_{fname}")
            broker_map[fname] = None if choice == "(None)" else choice

    with map_col2:
        st.markdown("**Murex File columns**")
        for fname in CANONICAL_FIELDS:
            options = ["(None)"] + list(murex_raw.columns)
            default = DEFAULT_MUREX_COLUMN_MAP.get(fname)
            idx = options.index(default) if default in options else 0
            choice = st.selectbox(f"{fname}", options, index=idx, key=f"m_{fname}")
            murex_map[fname] = None if choice == "(None)" else choice

    st.subheader("🎯 Matching Configuration")
    cfg_col1, cfg_col2, cfg_col3 = st.columns(3)
    with cfg_col1:
        id_prefixes_text = st.text_input(
            "ID prefixes to strip (comma-separated)",
            value=", ".join(ID_STRIP_PREFIXES),
            help="System tags like 'BFU:' or 'ICE:' in front of the Murex G.ID that should be "
                 "stripped before comparing to the broker's plain numeric ID.")
        id_strip_prefixes = [p.strip() for p in id_prefixes_text.split(",") if p.strip()]
    with cfg_col2:
        price_tol = st.number_input("Price tolerance", min_value=0.0, value=DEFAULT_PRICE_TOLERANCE,
                                     step=0.0001, format="%.4f")
        qty_tol = st.number_input("Quantity tolerance", min_value=0.0, value=DEFAULT_QTY_TOLERANCE, step=0.01)
    with cfg_col3:
        enable_fallback = st.checkbox("Enable fallback attribute matching for no-ID rows", value=True)
        fallback_keys = st.multiselect(
            "Fallback match/group fields", options=[f for f in CANONICAL_FIELDS if f not in ("trade_id", "link_id")],
            default=DEFAULT_GROUP_KEYS)
        enable_cross_id = st.checkbox(
            "Enable cross-ID aggregation (enhancement layer)",
            value=DEFAULT_ENABLE_CROSS_ID_AGGREGATION,
            help="When an ID-matched pair fails only on lots/quantity, looks for OTHER broker "
                 "rows (different IDs) matching the same trader and price that, combined, sum "
                 "exactly to the Murex quantity. Resolves cleanly-matched cases out of FAIL/"
                 "Missing into a separate 'Resolved by Cross-ID Aggregation' section. Never "
                 "guesses between multiple possible combinations - those are left as-is for "
                 "manual review. Uncheck to fully restore the original behaviour.")
        enable_missing_attr = st.checkbox(
            "Enable attribute aggregation for missing sets (Section 8)",
            value=DEFAULT_ENABLE_MISSING_ATTR_RESOLUTION,
            help="For Murex IDs with NO matching broker ID at all, searches broker rows in "
                 "the Missing-in-Murex pool by exact price, then trader, then commodity, and "
                 "accepts the match when exactly ONE unique lot-sum combination equals the "
                 "Murex quantity. Handles the case where Murex holds a parent/aggregated ID "
                 "while the broker sends granular per-execution IDs. Uncheck to disable.")

    config = ReconConfig(
        price_tolerance=price_tol,
        quantity_tolerance=qty_tol,
        group_keys=fallback_keys,
        enable_fallback_matching=enable_fallback,
        enable_cross_id_aggregation=enable_cross_id,
        enable_missing_attr_resolution=enable_missing_attr,
    )

    if st.button("▶️ Run Reconciliation", type="primary"):
        with st.spinner("Reconciling trades..."):
            broker_df = normalise(broker_raw, broker_map, "Broker", id_strip_prefixes)
            murex_df = normalise(murex_raw, murex_map, "Murex", id_strip_prefixes)
            result = reconcile(broker_df, murex_df, config)
            # Enhancement layer (Section 7) - runs strictly AFTER the normal
            # reconciliation above. Only ever moves items out of FAIL/Missing
            # into a separate, clearly-tagged resolved category. Never touches
            # an existing PASS. Fully reversible via the checkbox in the UI.
            _resolve_cross_id_aggregation(broker_df, murex_df, result, config)
            # Enhancement layer (Section 8) - runs after Section 7. Resolves
            # Murex orphans (Missing-in-Broker) against broker orphans
            # (Missing-in-Murex) by price + trader + commodity + lots-sum.
            # Same no-guessing policy. Fully reversible via checkbox.
            _resolve_missing_by_attribute_aggregation(broker_df, murex_df, result, config)

            with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
                build_report(result, config, tmp.name)
                report_bytes = open(tmp.name, "rb").read()

        st.success("Reconciliation complete!")

        n_pass = len(result.id_passed) + len(result.id_passed_aggregated)
        m1, m2, m3, m4, m5, m6, m7, m8 = st.columns(8)
        m1.metric("Broker IDs", result.broker_ids_count)
        m2.metric("Murex IDs", result.murex_ids_count)
        m3.metric("PASS", n_pass)
        m4.metric("FAIL", len(result.id_failed))
        m5.metric("Missing in Murex", len(result.missing_in_murex))
        m6.metric("Missing in Broker", len(result.missing_in_broker))
        m7.metric("Cross-ID Resolved", len(result.cross_id_resolved))
        m8.metric("Attr Resolved", len(result.missing_attr_resolved))

        n_cross_id_resolved_broker_ids = sum(len(r["broker_legs"]) for r in result.cross_id_resolved)
        n_missing_attr_resolved_broker_ids = sum(len(r["broker_legs"]) for r in result.missing_attr_resolved)
        pct = 100 * (n_pass + n_cross_id_resolved_broker_ids + n_missing_attr_resolved_broker_ids) / result.broker_ids_count if result.broker_ids_count else 0
        st.progress(min(1.0, pct / 100), text=f"{pct:.1f}% of broker trade IDs reconciled (PASS)")

        st.download_button(
            "⬇️ Download Reconciliation Report (Excel)",
            data=report_bytes,
            file_name="trade_reconciliation_report.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

        tabs = st.tabs(["PASS", "FAIL", "Missing in Murex", "Missing in Broker", "No-ID Exceptions",
                        "Cross-ID Resolved", "Attr Resolved (Missing)"])
        with tabs[0]:
            st.dataframe(pd.DataFrame([
                {"Link ID": r["link_id"], "Type": "Aggregated" if r["is_aggregated"] else "Single",
                 "Broker Qty": r["broker_qty_total"], "Murex Qty": r["murex_qty_total"],
                 "Price": r["broker_price"]}
                for r in result.id_passed + result.id_passed_aggregated
            ]))
        with tabs[1]:
            st.dataframe(pd.DataFrame([
                {"Link ID": r["link_id"], "Issues": "; ".join(i["reason"] for i in r["issues"]),
                 "Broker Qty": r["broker_qty_total"], "Murex Qty": r["murex_qty_total"],
                 "Note": r.get("cross_id_comment", "")}
                for r in result.id_failed
            ]))
        with tabs[2]:
            st.dataframe(pd.DataFrame([
                {"Link ID": e["link_id"], "Broker Qty Total": e["broker_qty_total"], "Legs": len(e["broker_legs"])}
                for e in result.missing_in_murex
            ]))
        with tabs[3]:
            st.dataframe(pd.DataFrame([
                {"Link ID": e["link_id"], "Murex Qty Total": e["murex_qty_total"],
                 "Legs": len(e["murex_legs"]), "Note": e.get("attr_match_comment", "")}
                for e in result.missing_in_broker
            ]))
        with tabs[4]:
            st.write(f"Fallback matched: {len(result.fallback_matched)}  |  "
                     f"Fallback aggregated: {len(result.fallback_aggregated)}  |  "
                     f"Unmatched broker: {len(result.no_id_unmatched_broker)}  |  "
                     f"Unmatched murex: {len(result.no_id_unmatched_murex)}")
        with tabs[5]:
            if result.cross_id_resolved:
                st.caption("Cases moved out of FAIL / Missing in Murex by the Cross-ID Aggregation "
                           "enhancement layer. See the 'Resolved - Cross-ID Aggregation' sheet in the "
                           "downloaded report for full leg-level detail.")
                st.dataframe(pd.DataFrame([
                    {"Link ID": r["link_id"], "Trader": r["matched_by"]["trader"],
                     "Price": r["matched_by"]["price"], "Broker Legs Combined": len(r["broker_legs"]),
                     "Broker Qty": r["broker_qty_total"], "Murex Qty": r["murex_qty_total"],
                     "Warnings": "; ".join(r.get("warning_notes", []))}
                    for r in result.cross_id_resolved
                ]))
            else:
                st.write("No cases were resolved by the cross-ID aggregation layer in this run.")
        with tabs[6]:
            if result.missing_attr_resolved:
                st.caption(
                    "Murex entries that had NO direct ID match in the broker file, but were "
                    "resolved by finding broker rows in the Missing-in-Murex pool whose lots "
                    "sum uniquely equals the Murex quantity (filtered by price, trader, commodity). "
                    "See the 'Resolved - Attr Aggregation' sheet in the downloaded report for "
                    "full leg-level detail."
                )
                st.dataframe(pd.DataFrame([
                    {"Murex Link ID": r["murex_link_id"],
                     "Trader": r["matched_by"]["trader"],
                     "Price": r["matched_by"]["price"],
                     "Commodity": r["matched_by"]["commodity"],
                     "Broker Legs Combined": len(r["broker_legs"]),
                     "Broker Qty": r["broker_qty_total"],
                     "Murex Qty": r["murex_qty_total"],
                     "Warnings": "; ".join(r.get("warning_notes", []))}
                    for r in result.missing_attr_resolved
                ]))
            else:
                st.write("No cases were resolved by the attribute aggregation (missing-set) layer in this run.")
else:
    st.info("Upload both the Broker file and the Murex extract to begin.")
